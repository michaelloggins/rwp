import base64
import io
import json
import logging
import re
from datetime import date, datetime
from pathlib import Path

import azure.functions as func

from audit import log_action, query_log
from config import get_settings
from models import RWPRequest
from synapse_client import SynapseClient

app = func.FunctionApp(http_auth_level=func.AuthLevel.ANONYMOUS)

logger = logging.getLogger("rwp")

# Column rename/reorder mapping for RWP (ePHI) report to match AR "Combo Billing" format.
# Maps pipeline column names → AR display names in desired output order.
_RWP_COLUMN_MAP = [
    ("TESTCODE", "Test Code"),
    ("DATEENTER", "Test Date"),
    ("COMPNAME", "Facility"),
    ("LAST_NAME", "Last Name"),
    ("FIRST_NAME", "First Name"),
    ("CLIENT_SAMPLE_ID", "Spec ID"),
    ("SPEC_SOURCE", "Spec Type"),
    ("DATE_COLLECTED", "Spec Date"),
    ("BIRTH_DATE", "DOB"),
    ("PRICE", "Cost"),
    ("CATEGORY", "Category"),
    ("FINAL", "FINAL"),
    ("NUMRES", "NUMRES"),
    ("RN2", "RN2"),
    ("PRICELISTID", "PRICELISTID"),
    ("ORDNO", "ORDNO"),
    ("EXTERNAL_ID", "EXTERNAL_ID"),
    ("STATE", "STATE"),
    ("CITY", "CITY"),
    ("ZIP", "ZIP"),
    ("CANCEL_STATUS", "CANCEL_STATUS"),
]


def _apply_rwp_column_map(
    rows: list[tuple], columns: list[str]
) -> tuple[list[tuple], list[str]]:
    """Reorder and rename columns for RWP report to match AR Combo Billing format.

    Returns (reordered_rows, new_column_names).
    """
    # Build index mapping: for each target column, find its position in source
    col_index = {c: i for i, c in enumerate(columns)}
    indices = []
    new_names = []
    for src_name, display_name in _RWP_COLUMN_MAP:
        if src_name in col_index:
            indices.append(col_index[src_name])
            new_names.append(display_name)

    reordered = [tuple(row[i] for i in indices) for row in rows]
    return reordered, new_names

# Resolve static file directories relative to this file
_BASE_DIR = Path(__file__).resolve().parent
_STATIC_DIR = _BASE_DIR / "static"
_ADDIN_DIR = _STATIC_DIR / "addin"

# MIME types for static file serving
_MIME_TYPES = {
    ".html": "text/html",
    ".js": "application/javascript",
    ".css": "text/css",
    ".xml": "application/xml",
    ".json": "application/json",
    ".png": "image/png",
    ".ico": "image/x-icon",
}


def _get_user_groups(req: func.HttpRequest) -> set[str]:
    """Extract Entra ID group Object IDs from the EasyAuth principal header."""
    principal_header = req.headers.get("X-MS-CLIENT-PRINCIPAL", "")
    if not principal_header:
        return set()

    try:
        decoded = base64.b64decode(principal_header)
        principal = json.loads(decoded)
    except (ValueError, json.JSONDecodeError):
        logger.warning("Failed to decode X-MS-CLIENT-PRINCIPAL header")
        return set()

    groups = set()
    for claim in principal.get("claims", []):
        if claim.get("typ") == "groups":
            groups.add(claim.get("val", ""))
    return groups


def _get_user_name(req: func.HttpRequest) -> str:
    """Extract the authenticated user's name/email from EasyAuth headers."""
    return req.headers.get("X-MS-CLIENT-PRINCIPAL-NAME", "anonymous")


def _check_report_authorization(
    req: func.HttpRequest, report_type: str
) -> func.HttpResponse | None:
    """Check if the user is authorized for the requested report type.

    Returns None if authorized, or a 403 HttpResponse if not.
    """
    settings = get_settings()

    # If group IDs aren't configured, skip authorization (local dev)
    if not settings.ephi_group_id:
        return None

    user_groups = _get_user_groups(req)
    user_name = _get_user_name(req)

    if report_type == "RWP":
        # Full ePHI report requires ePHI group membership
        if settings.ephi_group_id not in user_groups:
            logger.warning(
                "Access denied: user '%s' requested ePHI report without "
                "sg-rwp-ePHI-Users membership",
                user_name,
            )
            return func.HttpResponse(
                json.dumps({
                    "error": "Access denied. The full Results With Pricing report "
                    "contains ePHI and requires membership in sg-rwp-ePHI-Users. "
                    "Contact your administrator to request access."
                }),
                status_code=403,
                mimetype="application/json",
            )
    else:
        # CFO report requires either group
        if (
            settings.ephi_group_id not in user_groups
            and settings.cfo_group_id not in user_groups
        ):
            logger.warning(
                "Access denied: user '%s' requested CFO report without "
                "group membership",
                user_name,
            )
            return func.HttpResponse(
                json.dumps({
                    "error": "Access denied. You must be a member of "
                    "sg-rwp-ePHI-Users or sg-rwp-CFO-Users to access this report. "
                    "Contact your administrator to request access."
                }),
                status_code=403,
                mimetype="application/json",
            )

    return None


def _require_authenticated(req: func.HttpRequest) -> func.HttpResponse | None:
    """Defense-in-depth: reject requests with no EasyAuth principal.

    Even though EasyAuth should block unauthenticated requests at the
    platform level, this guard protects against misconfiguration
    (e.g. EasyAuth disabled or entraClientId not set).
    """
    if not req.headers.get("X-MS-CLIENT-PRINCIPAL"):
        return func.HttpResponse(
            json.dumps({"error": "Authentication required"}),
            status_code=401,
            mimetype="application/json",
        )
    return None


@app.route(route="api/results-with-pricing", methods=["GET"])
async def results_with_pricing(req: func.HttpRequest) -> func.HttpResponse:
    """HTTP trigger: query RWP or RWPCFO view by date range."""

    # Defense-in-depth: require authentication even if EasyAuth is misconfigured
    auth_err = _require_authenticated(req)
    if auth_err:
        return auth_err

    # Parse and validate parameters
    try:
        params = RWPRequest.from_request(req)
    except ValueError as e:
        return func.HttpResponse(
            json.dumps({"error": str(e)}),
            status_code=400,
            mimetype="application/json",
        )

    # Enforce group-based authorization
    auth_error = _check_report_authorization(req, params.report_type)
    if auth_error:
        return auth_error

    # Select view based on report type
    view_name = (
        "dbo.vw_ResultsWithPricingCFO"
        if params.report_type == "RWPCFO"
        else "dbo.vw_ResultsWithPricing"
    )

    # Build GROUP BY query — date filtering happens BEFORE aggregation (matches SP)
    # Individual result rows are filtered by date range, THEN grouped with MIN/MAX
    #
    # _partition_year and _partition_month filters enable Synapse OPENROWSET
    # filepath() partition elimination — only the year/month folders covering
    # the date range are scanned.
    partition_years = list(range(params.cal_date.year, params.cal_end.year + 1))
    year_placeholders = ", ".join("?" for _ in partition_years)

    # Build month list covering the date range
    partition_months = []
    y, m = params.cal_date.year, params.cal_date.month
    while (y, m) <= (params.cal_end.year, params.cal_end.month):
        partition_months.append(m)
        m += 1
        if m > 12:
            m = 1
            y += 1
    # Deduplicate months (e.g. Jan-Jan across two years → just [1])
    partition_months = list(set(partition_months))
    month_placeholders = ", ".join("?" for _ in partition_months)

    if params.report_type == "RWPCFO":
        select_cols = """
            MIN(v.DATE_COLLECTED) AS DATE_COLLECTED,
            MIN(v.EXTERNAL_ID) AS EXTERNAL_ID,
            MIN(v.SPEC_SOURCE) AS SPEC_SOURCE,"""
    else:
        select_cols = """
            MIN(v.BIRTH_DATE) AS BIRTH_DATE,
            MIN(v.CLIENT_SAMPLE_ID) AS CLIENT_SAMPLE_ID,
            MIN(v.DATE_COLLECTED) AS DATE_COLLECTED,
            MIN(v.EXTERNAL_ID) AS EXTERNAL_ID,
            MIN(v.FIRST_NAME) AS FIRST_NAME,
            MIN(v.LAST_NAME) AS LAST_NAME,
            MIN(v.SPEC_SOURCE) AS SPEC_SOURCE,"""

    query = (
        "SELECT"
        + select_cols
        + """
            MIN(v.TESTCODE) AS TESTCODE,
            MIN(v.CATEGORY) AS CATEGORY,
            v.COMPNAME,
            MAX(v.DATEENTER) AS DATEENTER,
            MIN(v.FINAL) AS FINAL,
            MIN(v.NUMRES) AS NUMRES,
            LTRIM(RTRIM(MIN(v.RN2))) AS RN2,
            MIN(v.PRICE) AS PRICE,
            MIN(v.PRICELISTID) AS PRICELISTID,
            v.ORDNO,
            MIN(v.STATE) AS STATE,
            RTRIM(MIN(v.City)) AS CITY,
            RTRIM(REPLACE(MIN(v.Zip), CHAR(10), '')) AS ZIP,
            MIN(v.CANCEL_STATUS) AS CANCEL_STATUS
        FROM """
        + view_name
        + """ v
        WHERE v._partition_year IN ("""
        + year_placeholders
        + """)
          AND v._partition_month IN ("""
        + month_placeholders
        + """)
          AND v.DATEENTER >= ?
          AND v.DATEENTER <= ?
          AND v.LatestResultDate <= ?
          AND NOT EXISTS (
              SELECT 1
              FROM """
        + view_name
        + """ sc
              WHERE sc._partition_year IN ("""
        + year_placeholders
        + """)
                AND sc._partition_month IN ("""
        + month_placeholders
        + """)
                AND sc.EXTERNAL_ID = v.EXTERNAL_ID
                AND sc.COMPNAME LIKE 'Santa Clara%'
                AND sc.SourceTestCode = 315
                AND sc.DATEENTER >= ?
                AND sc.DATEENTER <= ?
          )
        GROUP BY v.ORDNO, v.COMPNAME
        ORDER BY v.COMPNAME, v.ORDNO
    """
    )
    # Convert all params to strings — aioodbc/Synapse OPENROWSET handles
    # string-to-int implicit conversion, but mixed int/str param lists can
    # cause binding errors in some ODBC driver versions.
    query_params = [
        *[str(y) for y in partition_years],
        *[str(m) for m in partition_months],
        params.cal_date.isoformat(), params.cal_end.isoformat(), params.cal_end.isoformat(),
        *[str(y) for y in partition_years],
        *[str(m) for m in partition_months],
        params.cal_date.isoformat(), params.cal_end.isoformat(),
    ]

    logger.info(
        "Executing query: view=%s, years=%s, months=%s, date_range=%s to %s, "
        "param_count=%d, format=%s",
        view_name, partition_years, partition_months,
        params.cal_date, params.cal_end, len(query_params), params.output_format,
    )

    settings = get_settings()
    client = SynapseClient(settings)

    # Derive audit fields
    user_name = _get_user_name(req)
    action = "view" if params.output_format == "json" else "download"
    version = params.report_type

    try:
        rows, columns = await client.execute_query(query, query_params)
    except Exception:
        logger.exception("Failed to execute Synapse query")
        log_action(
            user=user_name,
            year=params.cal_date.year,
            month=params.cal_date.month,
            action=action,
            version=version,
            filename="screen" if params.output_format == "json" else "",
            row_count=0,
            status="error",
        )
        return func.HttpResponse(
            json.dumps({"error": "Internal server error querying Synapse"}),
            status_code=500,
            mimetype="application/json",
        )

    logger.info("Query returned %d rows, %d columns", len(rows), len(columns))

    # Apply column rename/reorder for RWP (ePHI) report
    if params.report_type == "RWP":
        rows, columns = _apply_rwp_column_map(rows, columns)

    # Return in requested format
    try:
        if params.output_format == "csv":
            response = _build_csv_response(rows, columns)
            filename = "results_with_pricing.csv"
        elif params.output_format == "xlsx":
            response = _build_xlsx_response(
                rows, columns, params.cal_date, params.cal_end, params.report_type
            )
            filename = f"ResultsWithPricing_{params.cal_date.isoformat()}_{params.cal_end.isoformat()}.xlsx"
        else:
            response = _build_json_response(rows, columns)
            filename = "screen"

        log_action(
            user=user_name,
            year=params.cal_date.year,
            month=params.cal_date.month,
            action=action,
            version=version,
            filename=filename,
            row_count=len(rows),
            status="success",
        )
        return response
    except Exception:
        logger.exception("Failed to build %s response", params.output_format)
        log_action(
            user=user_name,
            year=params.cal_date.year,
            month=params.cal_date.month,
            action=action,
            version=version,
            filename="",
            row_count=len(rows),
            status="error",
        )
        return func.HttpResponse(
            json.dumps({"error": f"Internal server error building {params.output_format} response"}),
            status_code=500,
            mimetype="application/json",
        )


def _build_json_response(
    rows: list[tuple], columns: list[str]
) -> func.HttpResponse:
    """Convert rows to JSON array of objects."""
    results = []
    for row in rows:
        record = {}
        for i, col in enumerate(columns):
            val = row[i]
            # Handle datetime serialization
            if hasattr(val, "isoformat"):
                val = val.isoformat()
            record[col] = val
        results.append(record)

    return func.HttpResponse(
        json.dumps(results, default=str),
        status_code=200,
        mimetype="application/json",
    )


def _build_csv_response(
    rows: list[tuple], columns: list[str]
) -> func.HttpResponse:
    """Convert rows to CSV with header row."""
    lines = [",".join(columns)]
    for row in rows:
        values = []
        for val in row:
            if val is None:
                values.append("")
            elif hasattr(val, "isoformat"):
                values.append(val.isoformat())
            else:
                s = str(val)
                if "," in s or '"' in s or "\n" in s:
                    s = '"' + s.replace('"', '""') + '"'
                values.append(s)
        lines.append(",".join(values))

    csv_content = "\n".join(lines)
    return func.HttpResponse(
        csv_content,
        status_code=200,
        mimetype="text/csv",
        headers={
            "Content-Disposition": "attachment; filename=results_with_pricing.csv"
        },
    )


def _build_xlsx_response(
    rows: list[tuple],
    columns: list[str],
    cal_date: date,
    cal_end: date,
    report_type: str = "RWPCFO",
) -> func.HttpResponse:
    """Convert rows to a styled Excel workbook.

    For RWP reports, creates 4 tabs matching the AR "Combo Billing" format:
      - "Combo Billing" (all rows)
      - "Clin {Mon}" (Category = 'Clinical')
      - "Vet {Mon}" (Category = 'Veterinary')
      - "VA Indy" (Facility LIKE 'VA Indianapolis%')
    For RWPCFO, creates a single "Results With Pricing" tab.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill, numbers
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    # Header styling
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    # Identify date and price columns for formatting
    date_cols = set()
    price_cols = set()
    for i, col in enumerate(columns):
        col_lower = col.lower()
        if "date" in col_lower or col_lower in ("caldate", "calend", "dob"):
            date_cols.add(i)
        if col_lower in ("price", "cost", "amount", "fee"):
            price_cols.add(i)

    def _write_sheet(ws, sheet_rows):
        """Write headers and data rows to a worksheet."""
        # Header row
        for col_idx, col_name in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        # Data rows
        for row_idx, row in enumerate(sheet_rows, 2):
            for col_idx, val in enumerate(row):
                cell = ws.cell(row=row_idx, column=col_idx + 1)
                if val is None:
                    cell.value = None
                elif isinstance(val, (datetime, date)):
                    cell.value = val
                    cell.number_format = "YYYY-MM-DD"
                elif col_idx in price_cols and isinstance(val, (int, float)):
                    cell.value = val
                    cell.number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
                else:
                    cell.value = val

        # Auto-size columns (sample first 100 rows + header)
        for col_idx, col_name in enumerate(columns, 1):
            max_len = len(str(col_name))
            sample = min(len(sheet_rows), 100)
            for r in range(sample):
                val = sheet_rows[r][col_idx - 1]
                if val is not None:
                    max_len = max(max_len, len(str(val)))
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 50)

        # Auto-filter and freeze panes
        if columns:
            ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{len(sheet_rows) + 1}"
        ws.freeze_panes = "A2"

    if report_type == "RWP":
        # Build column indices for filtering
        col_index = {c: i for i, c in enumerate(columns)}
        cat_idx = col_index.get("Category")
        facility_idx = col_index.get("Facility")

        # Month abbreviation from cal_date
        month_abbr = cal_date.strftime("%b")

        # Tab definitions: (tab_name, filter_func)
        tabs = [
            ("Combo Billing", lambda _row: True),
            (f"Clin {month_abbr}", lambda _row: cat_idx is not None and _row[cat_idx] == "Clinical"),
            (f"Vet {month_abbr}", lambda _row: cat_idx is not None and _row[cat_idx] == "Veterinary"),
            ("VA Indy", lambda _row: facility_idx is not None and isinstance(_row[facility_idx], str) and _row[facility_idx].startswith("VA Indianapolis")),
        ]

        for i, (tab_name, filter_fn) in enumerate(tabs):
            if i == 0:
                ws = wb.active
                ws.title = tab_name
            else:
                ws = wb.create_sheet(title=tab_name)
            filtered = [r for r in rows if filter_fn(r)]
            _write_sheet(ws, filtered)
    else:
        # Single tab for RWPCFO
        ws = wb.active
        ws.title = "Results With Pricing"
        _write_sheet(ws, rows)

    # Write to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"ResultsWithPricing_{cal_date.isoformat()}_{cal_end.isoformat()}.xlsx"

    return func.HttpResponse(
        output.getvalue(),
        status_code=200,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@app.route(route="api/audit-log", methods=["GET"])
async def audit_log(req: func.HttpRequest) -> func.HttpResponse:
    """HTTP trigger: query the audit log for a given month."""

    # Defense-in-depth: require authentication even if EasyAuth is misconfigured
    auth_err = _require_authenticated(req)
    if auth_err:
        return auth_err

    settings = get_settings()

    # Restrict to ePHI group members
    if settings.ephi_group_id:
        user_groups = _get_user_groups(req)
        if settings.ephi_group_id not in user_groups:
            return func.HttpResponse(
                json.dumps({"error": "Access denied. Audit log requires sg-rwp-ePHI-Users membership."}),
                status_code=403,
                mimetype="application/json",
            )

    month_param = req.params.get("Month", "")
    if not month_param:
        return func.HttpResponse(
            json.dumps({"error": "Month parameter is required (YYYY-MM)"}),
            status_code=400,
            mimetype="application/json",
        )

    # Basic format validation
    if not re.match(r"^\d{4}-\d{2}$", month_param):
        return func.HttpResponse(
            json.dumps({"error": "Month must be in YYYY-MM format"}),
            status_code=400,
            mimetype="application/json",
        )

    try:
        entries = query_log(month_param)
    except Exception:
        logger.exception("Failed to query audit log")
        return func.HttpResponse(
            json.dumps({"error": "Internal server error querying audit log"}),
            status_code=500,
            mimetype="application/json",
        )

    return func.HttpResponse(
        json.dumps(entries, default=str),
        status_code=200,
        mimetype="application/json",
    )


# --- Static file serving routes ----------------------------------------------


@app.route(route="", methods=["GET"], auth_level=func.AuthLevel.ANONYMOUS)
async def root_redirect(req: func.HttpRequest) -> func.HttpResponse:
    """Redirect root URL to /download."""
    return func.HttpResponse(status_code=301, headers={"Location": "/download"})


@app.route(route="download", methods=["GET"], auth_level=func.AuthLevel.ANONYMOUS)
async def download_page(req: func.HttpRequest) -> func.HttpResponse:
    """Serve the web download page."""
    return _serve_static_file(_STATIC_DIR / "index.html")


@app.route(route="addin/{*filepath}", methods=["GET"], auth_level=func.AuthLevel.ANONYMOUS)
async def addin_files(req: func.HttpRequest) -> func.HttpResponse:
    """Serve Excel Add-in static files."""
    filepath = req.route_params.get("filepath", "")
    if not filepath:
        filepath = "taskpane.html"
    return _serve_static_file(_ADDIN_DIR / filepath)


def _serve_static_file(file_path: Path) -> func.HttpResponse:
    """Read and return a static file with the correct MIME type."""
    # Prevent directory traversal
    try:
        file_path = file_path.resolve()
    except (ValueError, OSError):
        return func.HttpResponse("Not found", status_code=404)

    # Verify resolved path is within the allowed static directories
    if not (file_path.is_relative_to(_STATIC_DIR)):
        return func.HttpResponse("Not found", status_code=404)

    if not file_path.is_file():
        return func.HttpResponse("Not found", status_code=404)

    suffix = file_path.suffix.lower()
    mime = _MIME_TYPES.get(suffix, "application/octet-stream")

    content = file_path.read_bytes()
    return func.HttpResponse(content, status_code=200, mimetype=mime)
